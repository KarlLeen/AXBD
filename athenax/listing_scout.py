"""Lightweight listing-scout helpers: exclusion inventory, JSON parse, xlsx I/O.

Phase A of "Scout then Enrich 200": discover NEW projects with only
name / Website / Twitter (+ optional category). Phase B is handled by
scripts/enrich_listing.py + scripts/refill_team_serper.py (see
scripts/enrich_scout_batch.py).
"""
from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

import httpx

# Match W6 listing columns (data/athenax_listing_w6.xlsx).
W6_HEADERS = [
    "name",
    "category",
    "subcategory",
    "Website",
    "Twitter",
    "short_desc",
    "description",
    "stage",
    "founded",
    "Discord",
    "GitHub",
    "Docs",
    "team",
    "backers",
]

DEFAULT_SHEET = "Scout Batch"

SECTOR_ROTATION = [
    "AI & Agents",
    "Crypto",
    "Developer Tools",
    "Infrastructure",
    "RWA",
    "Biotech",
    "Robotics",
]

_VALID_CATEGORIES = {s.casefold(): s for s in SECTOR_ROTATION}

_URL_RE = re.compile(r"^https?://", re.I)
_HANDLE_RE = re.compile(r"^@?[A-Za-z0-9_]{1,15}$")


def casefold_name(name: str | None) -> str:
    return (name or "").strip().casefold()


def load_inventory(path: str | Path) -> dict:
    p = Path(path)
    if not p.exists():
        return {
            "approved_names": [],
            "pending_names": [],
            "union_unique_names": 0,
        }
    with open(p, encoding="utf-8") as f:
        data = json.load(f)
    return data if isinstance(data, dict) else {}


def names_from_inventory(inventory: dict) -> set[str]:
    """Casefolded union of approved + pending names from an inventory dict."""
    out: set[str] = set()
    for key in ("approved_names", "pending_names"):
        for n in inventory.get(key) or []:
            cf = casefold_name(str(n))
            if cf:
                out.add(cf)
    return out


def fetch_approved_names(
    url: str = "https://api.athenax.co/api/v1/product",
    limit: int = 800,
    timeout: float = 30.0,
) -> list[str]:
    """Fetch approved product names from the public AthenaX API."""
    resp = httpx.get(
        url,
        params={"status": "approved", "limit": limit},
        timeout=timeout,
    )
    resp.raise_for_status()
    data = resp.json()
    if isinstance(data, list):
        items = data
    elif isinstance(data, dict):
        items = (
            data.get("items")
            or data.get("data")
            or data.get("products")
            or data.get("results")
            or []
        )
    else:
        items = []
    names: list[str] = []
    seen: set[str] = set()
    for item in items:
        if not isinstance(item, dict):
            continue
        n = (item.get("name") or "").strip()
        cf = casefold_name(n)
        if n and cf not in seen:
            seen.add(cf)
            names.append(n)
    return sorted(names, key=str.casefold)


def refresh_approved_inventory(
    path: str | Path,
    api_url: str = "https://api.athenax.co/api/v1/product",
    limit: int = 800,
) -> dict:
    """Refresh approved_names from the public API; keep pending_names as-is."""
    path = Path(path)
    inv = load_inventory(path)
    approved = fetch_approved_names(url=api_url, limit=limit)
    pending = [str(n).strip() for n in (inv.get("pending_names") or []) if str(n).strip()]
    pending_cf = {casefold_name(n) for n in pending}
    approved_cf = {casefold_name(n) for n in approved}
    overlap = approved_cf & pending_cf
    union = approved_cf | pending_cf
    inv.update(
        {
            "source_approved": f"{api_url}?status=approved&limit={limit}",
            "approved_count": len(approved),
            "approved_unique_names": len(approved_cf),
            "pending_count": len(pending),
            "pending_unique_names": len(pending_cf),
            "overlap_names": len(overlap),
            "union_unique_names": len(union),
            "approved_names": approved,
            "pending_names": pending,
        }
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(inv, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return inv


def names_from_xlsx(
    path: str | Path,
    sheet: str | None = None,
    name_header: str = "name",
) -> set[str]:
    """Casefolded project names already present in an output workbook."""
    path = Path(path)
    if not path.exists():
        return set()
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            return set()
        header_l = [str(h).strip() if h is not None else "" for h in header]
        try:
            idx = header_l.index(name_header)
        except ValueError:
            return set()
        out: set[str] = set()
        for row in rows:
            if not row or idx >= len(row):
                continue
            cf = casefold_name(str(row[idx]) if row[idx] is not None else "")
            if cf:
                out.add(cf)
        return out
    finally:
        wb.close()


def build_exclude_set(
    inventory_path: str | Path,
    output_xlsx: str | Path | None = None,
    sheet: str | None = None,
    refresh_approved: bool = False,
) -> set[str]:
    """Load excluded names (inventory ∪ existing output), optionally refresh API."""
    if refresh_approved:
        inv = refresh_approved_inventory(inventory_path)
    else:
        inv = load_inventory(inventory_path)
    excluded = names_from_inventory(inv)
    if output_xlsx:
        excluded |= names_from_xlsx(output_xlsx, sheet=sheet)
    return excluded


def format_exclude_context(excluded: set[str], max_names: int = 1200) -> str:
    """Prompt block listing excluded names (sorted)."""
    if not excluded:
        return (
            "━━━ EXCLUSION LIST ━━━\n"
            "No excluded names loaded. Still avoid inventing projects.\n"
        )
    names = sorted(excluded)
    truncated = ""
    if len(names) > max_names:
        names = names[:max_names]
        truncated = f"\n(Showing first {max_names} of {len(excluded)} excluded names.)\n"
    lines = "\n".join(f"  - {n}" for n in names)
    return (
        "━━━ EXCLUSION LIST — DO NOT INCLUDE THESE ━━━\n"
        "These projects are already approved, pending, or already in this batch. "
        "Scout DIFFERENT projects. Compare case-insensitively.\n"
        f"{lines}{truncated}"
    )


def _balanced_extract(text: str, start: int) -> str | None:
    depth = 0
    in_string = False
    escape = False
    for i in range(start, len(text)):
        ch = text[i]
        if escape:
            escape = False
            continue
        if in_string:
            if ch == "\\":
                escape = True
            elif ch == '"':
                in_string = False
            continue
        if ch == '"':
            in_string = True
        elif ch in ("[", "{"):
            depth += 1
        elif ch in ("]", "}"):
            depth -= 1
            if depth == 0:
                return text[start : i + 1]
    return None


def _try_parse(s: str) -> Any:
    try:
        return json.loads(s)
    except json.JSONDecodeError:
        pass
    fixed = re.sub(r",(\s*[}\]])", r"\1", s)
    try:
        return json.loads(fixed)
    except json.JSONDecodeError:
        pass
    try:
        from json_repair import repair_json

        repaired = repair_json(s, return_objects=True)
        if isinstance(repaired, (list, dict)):
            return repaired
    except Exception:
        pass
    return None


def parse_scout_json(text: str) -> list[dict]:
    """Extract a JSON array of scout lead objects from crew raw output."""
    if not text or not str(text).strip():
        return []
    text = re.sub(r"```(?:json)?\s*", "", str(text)).strip().rstrip("`").strip()

    candidates: list[list] = []

    def _classify(result: Any) -> None:
        if isinstance(result, dict):
            # Single object wrapped — accept if it looks like a lead.
            if result.get("name") and (result.get("website") or result.get("Website")):
                candidates.append([result])
            return
        if isinstance(result, list) and result and isinstance(result[0], dict):
            first = result[0]
            if first.get("name") or first.get("website") or first.get("Website"):
                candidates.append(result)

    pos = 0
    while True:
        pos = text.find("[", pos)
        if pos == -1:
            break
        chunk = _balanced_extract(text, pos)
        if chunk:
            _classify(_try_parse(chunk))
            pos += len(chunk)
        else:
            pos += 1

    if not candidates:
        first = text.find("[")
        if first != -1:
            try:
                from json_repair import repair_json

                _classify(repair_json(text[first:], return_objects=True))
            except Exception:
                pass

    if not candidates:
        # Last resort: a lone object.
        pos = text.find("{")
        if pos != -1:
            chunk = _balanced_extract(text, pos)
            if chunk:
                _classify(_try_parse(chunk))

    if not candidates:
        return []
    return max(candidates, key=len)


def looks_like_website(url: str | None) -> bool:
    if not url or not isinstance(url, str):
        return False
    u = url.strip()
    if not _URL_RE.match(u):
        return False
    try:
        p = urlparse(u)
    except Exception:
        return False
    if p.scheme not in ("http", "https"):
        return False
    host = (p.netloc or "").lower()
    if not host or "." not in host:
        return False
    # Reject obvious non-project placeholders.
    bad = ("example.com", "localhost", "127.0.0.1", "yourwebsite", "changeme")
    return not any(b in host for b in bad)


def normalize_twitter(value: str | None) -> str:
    """Return a full https://x.com/... URL, or '' if missing/invalid."""
    if not value or not isinstance(value, str):
        return ""
    v = value.strip()
    if not v or v.lower() in ("not found", "none", "null", "n/a"):
        return ""
    if _URL_RE.match(v):
        try:
            p = urlparse(v)
        except Exception:
            return ""
        host = (p.netloc or "").lower()
        if "twitter.com" not in host and "x.com" not in host:
            return v  # keep other social URLs as-is only if already a URL
        path = (p.path or "").strip("/")
        handle = path.split("/")[0] if path else ""
        if handle and _HANDLE_RE.match(handle):
            return f"https://x.com/{handle.lstrip('@')}"
        return v
    handle = v.lstrip("@").split("/")[-1]
    if _HANDLE_RE.match(handle):
        return f"https://x.com/{handle}"
    return ""


def normalize_lead(raw: dict) -> dict | None:
    """Normalize a scout object to {name, website, twitter, category?}."""
    if not isinstance(raw, dict):
        return None
    name = (raw.get("name") or "").strip()
    website = (raw.get("website") or raw.get("Website") or raw.get("url") or "").strip()
    twitter_raw = raw.get("twitter") or raw.get("Twitter") or raw.get("twitter_url") or ""
    category_raw = (raw.get("category") or "").strip()

    if not name or not looks_like_website(website):
        return None

    twitter = normalize_twitter(str(twitter_raw) if twitter_raw is not None else "")
    category = ""
    if category_raw:
        category = _VALID_CATEGORIES.get(category_raw.casefold(), category_raw)

    return {
        "name": name,
        "website": website.rstrip("/"),
        "twitter": twitter,
        "category": category,
    }


def filter_scout_leads(
    leads: list[dict],
    excluded: set[str],
    seen_names: set[str] | None = None,
    seen_websites: set[str] | None = None,
) -> tuple[list[dict], dict[str, int]]:
    """Filter/normalize scout leads. Returns (kept, stats)."""
    seen_names = set(seen_names or ())
    seen_websites = set(seen_websites or ())
    stats = {
        "raw": len(leads),
        "kept": 0,
        "no_website": 0,
        "excluded": 0,
        "duplicate_name": 0,
        "duplicate_website": 0,
    }
    kept: list[dict] = []
    for raw in leads:
        norm = normalize_lead(raw)
        if norm is None:
            stats["no_website"] += 1
            continue
        cf = casefold_name(norm["name"])
        if cf in excluded or cf in seen_names:
            if cf in excluded:
                stats["excluded"] += 1
            else:
                stats["duplicate_name"] += 1
            continue
        web_key = norm["website"].casefold()
        if web_key in seen_websites:
            stats["duplicate_website"] += 1
            continue
        seen_names.add(cf)
        seen_websites.add(web_key)
        kept.append(norm)
    stats["kept"] = len(kept)
    return kept, stats


def lead_to_w6_row(lead: dict) -> list:
    """Map a normalized scout lead into a W6 row (other columns blank)."""
    row = {h: "" for h in W6_HEADERS}
    row["name"] = lead.get("name") or ""
    row["Website"] = lead.get("website") or ""
    row["Twitter"] = lead.get("twitter") or ""
    row["category"] = lead.get("category") or ""
    return [row[h] for h in W6_HEADERS]


def ensure_workbook(path: str | Path, sheet: str = DEFAULT_SHEET):
    """Create or open a W6-header workbook. Returns (wb, ws)."""
    import openpyxl

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        wb = openpyxl.load_workbook(path)
        if sheet in wb.sheetnames:
            ws = wb[sheet]
        else:
            ws = wb.create_sheet(sheet)
            ws.append(W6_HEADERS)
        # Ensure header row exists / matches.
        if ws.max_row < 1 or not any(c.value for c in ws[1]):
            if ws.max_row >= 1:
                for i, h in enumerate(W6_HEADERS, start=1):
                    ws.cell(1, i).value = h
            else:
                ws.append(W6_HEADERS)
        return wb, ws

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(W6_HEADERS)
    return wb, ws


def count_xlsx_rows(path: str | Path, sheet: str | None = None) -> int:
    return len(names_from_xlsx(path, sheet=sheet))


def append_leads_to_xlsx(
    path: str | Path,
    leads: list[dict],
    sheet: str = DEFAULT_SHEET,
) -> int:
    """Append normalized leads to the workbook (dedup by name). Returns added count."""
    path = Path(path)
    existing = names_from_xlsx(path, sheet=sheet) if path.exists() else set()
    wb, ws = ensure_workbook(path, sheet=sheet)
    added = 0
    for lead in leads:
        cf = casefold_name(lead.get("name"))
        if not cf or cf in existing:
            continue
        ws.append(lead_to_w6_row(lead))
        existing.add(cf)
        added += 1
    wb.save(path)
    wb.close()
    return added


def sector_hint_for_round(round_index: int, override: str = "") -> str:
    if override and override.strip():
        return override.strip()
    return SECTOR_ROTATION[round_index % len(SECTOR_ROTATION)]
