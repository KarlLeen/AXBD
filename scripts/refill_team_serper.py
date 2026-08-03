"""Serper-assisted team LinkedIn refill — same strict match rules, no relaxation.

Two modes (both keep the hard gates from refill_team_linkedin.py):
  1. Named members missing LinkedIn:
       Serper '"Name" Company site:linkedin.com/in'
       → ConnectSafely people_search '"Name" Company'
       → accept only if name matches AND company in headline/current_company
       → if Serper surfaced /in/ URLs, prefer a people_search hit whose URL is in that set
         (still must pass name+company; never accept Serper URL alone)
  2. Empty team rows (--fill-empty-teams):
       Serper 'Company founder OR co-founder'
       → parse candidate names from titles/snippets
       → people_search each '"Name" Company' with the same strict gates
       → append Name | title |  | linkedin

No profile API. Sleep applies to ConnectSafely people_search only.

Usage:
    uv run python scripts/refill_team_serper.py \\
        --input data/athenax_listing_w6_team_work.xlsx \\
        --sheet "Week 6 (Huge list)" \\
        --output data/athenax_listing_w6_team_work.xlsx \\
        --start 0 --limit 40 --sleep 10
"""
from __future__ import annotations

import argparse
import difflib
import os
import re
import sys
import time
import unicodedata
from pathlib import Path

_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(_ROOT))
from dotenv import load_dotenv

load_dotenv(_ROOT / ".env")

import httpx
import openpyxl

_BASE = "https://api.connectsafely.ai/linkedin"
_LI_IN = re.compile(
    r"https?://(?:(?:www|[a-z]{2})\.)?linkedin\.com/in/([a-zA-Z0-9\-_%]+)/?",
    re.I,
)


def _cs_headers() -> dict:
    key = os.getenv("CONNECTSAFELY_API_KEY", "")
    if not key:
        sys.exit("CONNECTSAFELY_API_KEY not set")
    return {"Authorization": f"Bearer {key}", "Content-Type": "application/json"}


def _serper_key() -> str:
    key = os.getenv("SERPER_API_KEY", "")
    if not key:
        sys.exit("SERPER_API_KEY not set")
    return key


def _fold(s: str) -> str:
    s = unicodedata.normalize("NFKD", s or "")
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z ]", "", s.lower())


def _name_matches(stated: str, real: str | None) -> bool:
    if not stated or not real:
        return False
    a = _fold(stated).split()
    b = _fold(real).split()
    if not a or not b:
        return False
    if difflib.SequenceMatcher(None, a[-1], b[-1]).ratio() < 0.8:
        return False
    return difflib.SequenceMatcher(None, " ".join(a), " ".join(b)).ratio() >= 0.55


def _company_mentioned(company: str, *fields: str) -> bool:
    blob = " ".join(f or "" for f in fields).lower()
    if not blob:
        return False
    raw = re.sub(r"\([^)]*\)", " ", company).strip()
    tokens = [t for t in re.sub(r"[^a-z0-9 ]", " ", raw.lower()).split() if len(t) >= 3]
    stop = {"the", "and", "inc", "ltd", "llc", "labs", "lab", "ai", "web", "app", "for"}
    tokens = [t for t in tokens if t not in stop]
    if not tokens:
        needle = re.sub(r"[^a-z0-9]", "", raw.lower())
        return len(needle) >= 3 and needle in re.sub(r"[^a-z0-9]", "", blob)
    tokens.sort(key=len, reverse=True)
    return any(t in blob for t in tokens[:2])


def _parse_team(cell: str) -> list[list[str]]:
    if not cell:
        return []
    people = []
    for chunk in str(cell).split(" ; "):
        parts = [p.strip() for p in chunk.split(" | ")]
        while len(parts) < 4:
            parts.append("")
        people.append(parts[:4])
    return people


def _render_team(people: list[list[str]]) -> str:
    out = []
    for name, title, bio, linkedin in people:
        parts = [p for p in (name, title, bio, linkedin) if p]
        if parts:
            out.append(" | ".join(parts))
    return " ; ".join(out)


def _is_rate_limit(status: int, body: str) -> bool:
    if status == 429:
        return True
    if status < 400:
        return False
    low = (body or "").lower()
    return (
        "rate limit" in low
        or "rate_limit" in low
        or "daily limit" in low
        or "daily quota" in low
        or "quota exceeded" in low
        or "too many requests" in low
        or ("limit" in low and ("exceed" in low or "reached" in low or "month" in low))
    )


def _is_auth_fail(status: int, body: str) -> bool:
    if status < 400:
        return False
    low = (body or "").lower()
    return "authentication credentials" in low or "linkedin authentication" in low


def _normalize_li_url(url: str) -> str:
    m = _LI_IN.search(url or "")
    if not m:
        return ""
    return f"https://www.linkedin.com/in/{m.group(1).strip('/')}"


def serper_search(query: str, num: int = 8) -> tuple[list[dict], str | None]:
    try:
        r = httpx.post(
            "https://google.serper.dev/search",
            headers={"X-API-KEY": _serper_key(), "Content-Type": "application/json"},
            json={"q": query, "num": max(1, min(num, 10))},
            timeout=20,
        )
    except httpx.RequestError as e:
        return [], f"serper request error: {e}"
    if r.status_code >= 400:
        return [], f"serper HTTP {r.status_code}: {r.text[:200]}"
    data = r.json() if r.content else {}
    out = []
    for item in data.get("organic") or []:
        out.append({
            "title": item.get("title") or "",
            "url": item.get("link") or "",
            "snippet": item.get("snippet") or "",
        })
    return out, None


def people_search(keywords: str, count: int = 10) -> tuple[list[dict], str | None]:
    try:
        r = httpx.post(
            f"{_BASE}/search/people",
            headers=_cs_headers(),
            json={"keywords": keywords, "limit": count},
            timeout=45,
        )
    except httpx.RequestError as e:
        return [], f"request error: {e}"
    if _is_rate_limit(r.status_code, r.text):
        return [], "RATE_LIMIT"
    if _is_auth_fail(r.status_code, r.text):
        return [], "AUTH_FAIL"
    if r.status_code >= 400:
        return [], f"HTTP {r.status_code}: {r.text[:200]}"
    data = r.json() if r.content else {}
    results = []
    for p in data.get("people") or data.get("results") or []:
        url = (p.get("profileUrl") or "").strip()
        if not url:
            pid = (p.get("profileId") or p.get("publicIdentifier") or "").strip()
            if pid and "/" not in pid:
                url = f"https://www.linkedin.com/in/{pid}"
        results.append({
            "name": f"{p.get('firstName', '')} {p.get('lastName', '')}".strip(),
            "headline": p.get("headline") or "",
            "linkedin_profile": _normalize_li_url(url) or url,
            "current_company": p.get("currentCompany") or p.get("currentPosition") or "",
        })
    return results, None


def _serper_li_candidates(company: str, person_name: str, hits: list[dict]) -> set[str]:
    """LinkedIn /in/ URLs from Serper that mention company and look like this person."""
    out: set[str] = set()
    for h in hits:
        url = _normalize_li_url(h.get("url") or "")
        if not url:
            continue
        blob = f"{h.get('title', '')} {h.get('snippet', '')}"
        if not _company_mentioned(company, blob):
            continue
        # Title usually starts with the person's name on LinkedIn SERP cards
        title = h.get("title") or ""
        title_name = re.split(r"\s*[-–|•]\s*", title, maxsplit=1)[0].strip()
        title_name = re.sub(r"\s*\|\s*LinkedIn.*$", "", title_name, flags=re.I).strip()
        if _name_matches(person_name, title_name) or _name_matches(person_name, title):
            out.add(url)
    return out


def _pick_match(
    person_name: str,
    company: str,
    hits: list[dict],
    prefer_urls: set[str] | None = None,
) -> tuple[str | None, str]:
    """Return (url, title_from_headline) under strict name+company rules."""
    ranked = []
    for h in hits:
        url = _normalize_li_url(h.get("linkedin_profile") or "")
        if not url:
            continue
        if not _name_matches(person_name, h.get("name") or ""):
            continue
        if not _company_mentioned(company, h.get("headline") or "", h.get("current_company") or ""):
            continue
        prefer = 0 if (prefer_urls and url in prefer_urls) else 1
        ranked.append((prefer, url, h.get("headline") or ""))
    if not ranked:
        return None, ""
    ranked.sort(key=lambda x: x[0])
    # If Serper gave candidates, require an overlap — do not relax to unrelated people_search hits
    if prefer_urls:
        preferred = [x for x in ranked if x[0] == 0]
        if not preferred:
            return None, ""
        return preferred[0][1], preferred[0][2]
    return ranked[0][1], ranked[0][2]


_NAME_FROM_TEXT = re.compile(
    r"\b([A-Z][a-z]+(?:\s+[A-Z][a-z]+){1,2})\b"
)


def _names_from_serper(company: str, hits: list[dict]) -> list[str]:
    """Conservative founder-name harvest from Serper titles/snippets."""
    found: list[str] = []
    seen: set[str] = set()
    for h in hits:
        blob = f"{h.get('title', '')} {h.get('snippet', '')}"
        if not _company_mentioned(company, blob):
            continue
        low = blob.lower()
        if not any(k in low for k in ("founder", "co-founder", "cofounder", "ceo")):
            continue
        for m in _NAME_FROM_TEXT.finditer(blob):
            name = m.group(1).strip()
            # Skip company-like / junk tokens
            if _company_mentioned(company, name):
                continue
            if name.lower() in {"linkedin", "twitter", "github", "about"}:
                continue
            key = _fold(name)
            if key in seen or len(key.split()) < 2:
                continue
            seen.add(key)
            found.append(name)
            if len(found) >= 5:
                return found
    return found


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--sheet", required=True)
    ap.add_argument("--output", required=True)
    ap.add_argument("--start", type=int, default=0)
    ap.add_argument("--limit", type=int, default=0)
    ap.add_argument("--sleep", type=float, default=10.0,
                    help="Seconds between ConnectSafely people_search calls")
    ap.add_argument("--serper-sleep", type=float, default=1.0)
    ap.add_argument(
        "--fill-empty-teams",
        action="store_true",
        help="Also try to discover founders for rows with empty team via Serper",
    )
    ap.add_argument(
        "--require-serper-url",
        action="store_true",
        default=True,
        help="For named members, Serper must surface a matching /in/ URL before people_search can fill (default on)",
    )
    ap.add_argument("--no-require-serper-url", action="store_false", dest="require_serper_url")
    args = ap.parse_args()

    out_path = Path(args.output)
    src = out_path if out_path.exists() else Path(args.input)
    wb = openpyxl.load_workbook(src)
    ws = wb[args.sheet]
    header = [c.value for c in ws[1]]
    col = {name: i + 1 for i, name in enumerate(header) if name}
    if "team" not in col or "name" not in col:
        sys.exit("Sheet missing 'name' or 'team' column")

    data_rows = [r for r in range(2, ws.max_row + 1) if ws.cell(r, col["name"]).value]
    batch = data_rows[args.start: args.start + args.limit] if args.limit else data_rows[args.start:]

    serper_calls = cs_calls = filled = no_hit = skipped = errors = empty_filled = 0
    print(
        f"Serper→people_search refill for {len(batch)} rows "
        f"(start={args.start}, sleep={args.sleep}s, require_serper_url={args.require_serper_url}, "
        f"fill_empty={args.fill_empty_teams})...",
        file=sys.stderr,
    )

    for i, r in enumerate(batch, 1):
        company = str(ws.cell(r, col["name"]).value).strip()
        cell = ws.cell(r, col["team"]).value
        people = _parse_team(cell)
        changed = False

        # ── empty team: discover names via Serper, then strict people_search ──
        if not people and args.fill_empty_teams:
            q = f'{company} founder OR co-founder OR CEO'
            hits, err = serper_search(q, num=8)
            serper_calls += 1
            time.sleep(args.serper_sleep)
            if err:
                errors += 1
                print(f"  ! row {r} empty-team serper: {err}", file=sys.stderr, flush=True)
            else:
                for cand_name in _names_from_serper(company, hits):
                    kw = f'"{cand_name}" {company}'
                    ps, perr = people_search(kw)
                    cs_calls += 1
                    if perr == "RATE_LIMIT":
                        print(f"\nRATE_LIMIT at row {r} ({company}) — stopping.", file=sys.stderr)
                        print(f"Resume with: --start {args.start + i - 1}", file=sys.stderr)
                        wb.save(out_path)
                        return
                    if perr == "AUTH_FAIL":
                        print(f"\nAUTH_FAIL at row {r} ({company}) — stopping.", file=sys.stderr)
                        print(f"Resume with: --start {args.start + i - 1}", file=sys.stderr)
                        wb.save(out_path)
                        return
                    if perr:
                        errors += 1
                        print(f"  ! row {r} {cand_name!r}: {perr}", file=sys.stderr, flush=True)
                        time.sleep(args.sleep)
                        continue
                    url, headline = _pick_match(cand_name, company, ps, prefer_urls=None)
                    time.sleep(args.sleep)
                    if not url:
                        no_hit += 1
                        continue
                    hl = (headline or "").lower()
                    if "co-founder" in hl or "cofounder" in hl:
                        title = "Co-Founder"
                    elif "founder" in hl:
                        title = "Founder"
                    elif re.search(r"\bceo\b", hl):
                        title = "CEO"
                    elif re.search(r"\bcto\b", hl):
                        title = "CTO"
                    else:
                        title = ""
                    people.append([cand_name, title, "", url])
                    empty_filled += 1
                    changed = True
                    print(f"  ✓ row {r}: NEW {cand_name!r} @ {company} -> {url}",
                          file=sys.stderr, flush=True)

        # ── named members missing LinkedIn ──
        for person in people:
            name, title, bio, linkedin = person
            if not name:
                continue
            if linkedin and "linkedin.com/in/" in linkedin.lower():
                skipped += 1
                continue

            sq = f'"{name}" {company} site:linkedin.com/in'
            shits, serr = serper_search(sq, num=8)
            serper_calls += 1
            time.sleep(args.serper_sleep)
            if serr:
                errors += 1
                print(f"  ! row {r} serper {name!r}: {serr}", file=sys.stderr, flush=True)
                continue

            prefer = _serper_li_candidates(company, name, shits)
            if args.require_serper_url and not prefer:
                no_hit += 1
                continue

            kw = f'"{name}" {company}'
            ps, perr = people_search(kw)
            cs_calls += 1
            if perr == "RATE_LIMIT":
                print(f"\nRATE_LIMIT at row {r} ({company}) — stopping.", file=sys.stderr)
                print(f"Resume with: --start {args.start + i - 1}", file=sys.stderr)
                wb.save(out_path)
                return
            if perr == "AUTH_FAIL":
                print(f"\nAUTH_FAIL at row {r} ({company}) — stopping.", file=sys.stderr)
                print(f"Resume with: --start {args.start + i - 1}", file=sys.stderr)
                wb.save(out_path)
                return
            if perr:
                errors += 1
                print(f"  ! row {r} {name!r}: {perr}", file=sys.stderr, flush=True)
                time.sleep(args.sleep)
                continue

            url, _hl = _pick_match(
                name, company, ps,
                prefer_urls=prefer if args.require_serper_url else (prefer or None),
            )
            time.sleep(args.sleep)
            if not url:
                no_hit += 1
                continue
            person[3] = url
            filled += 1
            changed = True
            print(f"  ✓ row {r}: {name!r} @ {company} -> {url}", file=sys.stderr, flush=True)

        if changed:
            ws.cell(r, col["team"]).value = _render_team(people)
            wb.save(out_path)

        if i % 10 == 0:
            print(
                f"[{i}/{len(batch)}] serper={serper_calls} cs={cs_calls} "
                f"filled={filled} empty_new={empty_filled} no_hit={no_hit} errors={errors}",
                file=sys.stderr, flush=True,
            )

    wb.save(out_path)
    print(
        f"\nDone. serper={serper_calls} cs={cs_calls} filled={filled} "
        f"empty_new={empty_filled} no_hit={no_hit} skipped_has_url={skipped} errors={errors}",
        file=sys.stderr,
    )
    print(f"Saved {out_path}", file=sys.stderr)


if __name__ == "__main__":
    main()
