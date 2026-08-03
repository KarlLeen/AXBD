"""Phase A — Lightweight listing scout batch (~200 NEW projects).

Loops a lightweight CrewAI scout that returns only name / Website / Twitter
(+ optional category), filters against the admin exclusion inventory + rows
already in the output workbook, and appends W6-header rows (other columns blank).

Crash-safe: saves the .xlsx after every round. Resume: re-run with the same
--output; existing names are treated as excluded.

Does NOT push to api.athenax.co. Phase B enrichment:
    uv run python scripts/enrich_scout_batch.py \\
        --input data/scout_batch_200.xlsx \\
        --output data/scout_batch_200_enriched.xlsx

Usage:
    uv run python scripts/scout_listing_batch.py \\
        --target 200 \\
        --output data/scout_batch_200.xlsx \\
        --exclude data/admin_products_inventory.json \\
        --max-rounds 40 \\
        --per-round 20

    # Optional: refresh approved names from the public API before scouting
    uv run python scripts/scout_listing_batch.py --refresh-approved --target 200

    # Smoke (1 round):
    uv run python scripts/scout_listing_batch.py --target 20 --max-rounds 1
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(_ROOT))

from dotenv import load_dotenv

load_dotenv(_ROOT / ".env")

from athenax.crew import build_listing_scout_crew
from athenax.listing_scout import (
    DEFAULT_SHEET,
    append_leads_to_xlsx,
    build_exclude_set,
    count_xlsx_rows,
    filter_scout_leads,
    format_exclude_context,
    parse_scout_json,
    sector_hint_for_round,
)


def _progress(msg: str) -> None:
    print(msg, file=sys.stderr, flush=True)


def main() -> None:
    ap = argparse.ArgumentParser(
        description="Phase A: lightweight scout → W6 xlsx (name/Website/Twitter only)",
    )
    ap.add_argument("--target", type=int, default=200, help="Target NEW rows to collect")
    ap.add_argument(
        "--output",
        default="data/scout_batch_200.xlsx",
        help="Output workbook path (resumable)",
    )
    ap.add_argument(
        "--exclude",
        default="data/admin_products_inventory.json",
        help="Admin products inventory JSON (approved + pending names)",
    )
    ap.add_argument("--max-rounds", type=int, default=40, help="Max scout kickoffs")
    ap.add_argument(
        "--per-round",
        type=int,
        default=20,
        help="Hint for leads per round (prompt target; not a hard parser cap)",
    )
    ap.add_argument(
        "--sheet",
        default=DEFAULT_SHEET,
        help=f"Sheet name in the workbook (default: {DEFAULT_SHEET})",
    )
    ap.add_argument(
        "--sector-hint",
        default="",
        help="Optional fixed sector focus for every round (default: rotate)",
    )
    ap.add_argument(
        "--refresh-approved",
        action="store_true",
        help="Refresh approved_names in --exclude from the public AthenaX API",
    )
    ap.add_argument(
        "--dry-run-parse",
        default="",
        help="Offline: parse this text file as scout output and exit (no crew)",
    )
    args = ap.parse_args()

    out_path = Path(args.output)
    exclude_path = Path(args.exclude)

    if args.dry_run_parse:
        text = Path(args.dry_run_parse).read_text(encoding="utf-8")
        leads = parse_scout_json(text)
        excluded = build_exclude_set(
            exclude_path, output_xlsx=out_path, sheet=args.sheet, refresh_approved=False
        )
        kept, stats = filter_scout_leads(leads, excluded)
        _progress(f"dry-run-parse: raw={stats['raw']} kept={stats['kept']} stats={stats}")
        for lead in kept[: args.per_round]:
            _progress(f"  + {lead['name']} | {lead['website']} | {lead['twitter']}")
        if kept:
            n = append_leads_to_xlsx(out_path, kept, sheet=args.sheet)
            _progress(f"appended {n} → {out_path} (total {count_xlsx_rows(out_path, args.sheet)})")
        return

    # Initial exclude set (inventory ∪ existing output).
    excluded = build_exclude_set(
        exclude_path,
        output_xlsx=out_path,
        sheet=args.sheet,
        refresh_approved=args.refresh_approved,
    )
    existing_count = count_xlsx_rows(out_path, args.sheet) if out_path.exists() else 0
    _progress(
        f"Scout listing batch: target={args.target} max_rounds={args.max_rounds} "
        f"per_round≈{args.per_round}"
    )
    _progress(
        f"  exclude names loaded: {len(excluded)} "
        f"(inventory + {existing_count} already in {out_path})"
    )
    if existing_count >= args.target:
        _progress(f"Already at target ({existing_count} ≥ {args.target}). Nothing to do.")
        return

    seen_websites: set[str] = set()
    # Seed website dedup from existing workbook if present.
    if out_path.exists():
        import openpyxl

        wb = openpyxl.load_workbook(out_path, read_only=True, data_only=True)
        try:
            ws = wb[args.sheet] if args.sheet in wb.sheetnames else wb[wb.sheetnames[0]]
            rows = ws.iter_rows(values_only=True)
            header = next(rows, None)
            if header:
                h = [str(x).strip() if x is not None else "" for x in header]
                try:
                    wi = h.index("Website")
                except ValueError:
                    wi = -1
                if wi >= 0:
                    for row in rows:
                        if row and wi < len(row) and row[wi]:
                            seen_websites.add(str(row[wi]).strip().rstrip("/").casefold())
        finally:
            wb.close()

    total = existing_count
    for round_i in range(args.max_rounds):
        if total >= args.target:
            break

        # Rebuild exclude each round so we never re-ask for names we just saved.
        excluded = build_exclude_set(
            exclude_path,
            output_xlsx=out_path,
            sheet=args.sheet,
            refresh_approved=False,
        )
        seen_names = set(excluded)
        sector = sector_hint_for_round(round_i, args.sector_hint)
        exclude_context = format_exclude_context(excluded)
        need = args.target - total
        _progress(
            f"\n── Round {round_i + 1}/{args.max_rounds} ── "
            f"have={total}/{args.target} need={need} sector_focus={sector!r}"
        )

        crew = build_listing_scout_crew(
            exclude_context=exclude_context,
            sector_hint=sector,
        )
        try:
            result = crew.kickoff()
        except Exception as exc:
            _progress(f"  ! kickoff failed: {exc}")
            continue

        tasks_output = getattr(result, "tasks_output", [])
        raw_text = tasks_output[0].raw if tasks_output else str(result)
        leads = parse_scout_json(raw_text)
        kept, stats = filter_scout_leads(
            leads,
            excluded=excluded,
            seen_names=seen_names,
            seen_websites=seen_websites,
        )
        # Cap to remaining target (and soft per-round hint).
        cap = min(args.per_round, need)
        if len(kept) > cap:
            kept = kept[:cap]

        for lead in kept:
            seen_websites.add(lead["website"].casefold())

        added = append_leads_to_xlsx(out_path, kept, sheet=args.sheet) if kept else 0
        total = count_xlsx_rows(out_path, args.sheet)
        _progress(
            f"  parse={stats['raw']} kept_filtered={stats['kept']} "
            f"no_website={stats['no_website']} excluded={stats['excluded']} "
            f"dup_name={stats['duplicate_name']} dup_web={stats['duplicate_website']}"
        )
        _progress(f"  appended={added} → total={total}/{args.target} ({out_path})")
        for lead in kept:
            _progress(f"    + {lead['name']} | {lead['website']} | {lead['twitter'] or '—'}")

        if added == 0 and stats["raw"] == 0:
            _progress("  ! empty/unparseable scout output — continuing to next round")

    _progress(f"\nDone. {total} leads in {out_path} (target {args.target}).")
    if total < args.target:
        _progress("Target not reached — re-run to resume, or raise --max-rounds.")
    _progress(
        "Phase B (enrich, human review — no auto-push):\n"
        f"  uv run python scripts/enrich_scout_batch.py "
        f"--input {out_path} --output {out_path.with_name(out_path.stem + '_enriched.xlsx')}"
    )


if __name__ == "__main__":
    main()
