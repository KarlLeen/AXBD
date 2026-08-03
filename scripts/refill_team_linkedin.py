"""Deterministic (non-LLM) refill of missing team LinkedIn URLs via ConnectSafely.

For each team member who has a name but no LinkedIn URL, search
`"<Full Name>" <Company>` and only accept a hit when:
  1. the result name matches the stated name (same rules as verify_team_linkedin), and
  2. the company appears in headline or current_company.

Optionally confirms via the profile endpoint (name check again). Crash-safe:
saves after every changed row; stops cleanly on ConnectSafely rate limit.

Usage:
    uv run python scripts/refill_team_linkedin.py \\
        --input data/athenax_listing_w6_enriched.xlsx \\
        --sheet "Week 6 (Huge list)" \\
        --output data/athenax_listing_w6_enriched.xlsx
"""
from __future__ import annotations

import argparse
import difflib
import os
import re
import sys
import time
from pathlib import Path

_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(_ROOT))
from dotenv import load_dotenv

load_dotenv(_ROOT / ".env")

import httpx
import openpyxl

_BASE = "https://api.connectsafely.ai/linkedin"


def _headers() -> dict:
    key = os.getenv("CONNECTSAFELY_API_KEY", "")
    if not key:
        sys.exit("CONNECTSAFELY_API_KEY not set")
    return {"Authorization": f"Bearer {key}", "Content-Type": "application/json"}


def _name_matches(stated: str, real: str) -> bool:
    a = re.sub(r"[^a-z ]", "", stated.lower()).split()
    b = re.sub(r"[^a-z ]", "", real.lower()).split()
    if not a or not b:
        return False
    if difflib.SequenceMatcher(None, a[-1], b[-1]).ratio() < 0.8:
        return False
    return difflib.SequenceMatcher(None, " ".join(a), " ".join(b)).ratio() >= 0.55


def _company_mentioned(company: str, *fields: str) -> bool:
    """Require company token(s) to appear in headline/current_company."""
    blob = " ".join(f or "" for f in fields).lower()
    if not blob:
        return False
    # Strip parenthetical aliases: "Atlas (under thirdweb)" -> prefer longest meaningful token
    raw = re.sub(r"\([^)]*\)", " ", company).strip()
    tokens = [t for t in re.sub(r"[^a-z0-9 ]", " ", raw.lower()).split() if len(t) >= 3]
    # Drop ultra-generic words that create false positives
    stop = {"the", "and", "inc", "ltd", "llc", "labs", "lab", "ai", "web", "app", "for"}
    tokens = [t for t in tokens if t not in stop]
    if not tokens:
        # Fall back to full cleaned string if all tokens were short/stop
        needle = re.sub(r"[^a-z0-9]", "", raw.lower())
        return len(needle) >= 3 and needle in re.sub(r"[^a-z0-9]", "", blob)
    # Prefer the longest token (usually the distinctive brand)
    tokens.sort(key=len, reverse=True)
    return any(t in blob for t in tokens[:2])


def _parse_team(cell: str) -> list[list[str]]:
    if not cell:
        return []
    people = []
    for chunk in str(cell).split(" ; "):
        parts = [p.strip() for p in chunk.split(" | ")]
        while len(parts) < 4:
            parts.append("")
        # Keep only first 4 fields (ignore trailing twitter if present)
        people.append(parts[:4])
    return people


def _render_team(people: list[list[str]]) -> str:
    out = []
    for name, title, bio, linkedin in people:
        parts = [p for p in (name, title, bio, linkedin) if p]
        if parts:
            out.append(" | ".join(parts))
    return " ; ".join(out)


def _is_rate_limit(status: int, body: str) -> bool:
    """Only treat real quota/rate responses as RATE_LIMIT — never scan 200 bodies
    (headlines often contain words like 'daily' / 'limit')."""
    if status == 429:
        return True
    if status < 400:
        return False
    low = (body or "").lower()
    return (
        "rate limit" in low
        or "rate_limit" in low
        or "daily limit" in low
        or "daily quota" in low
        or "quota exceeded" in low
        or "too many requests" in low
        or ("limit" in low and ("exceed" in low or "reached" in low or "month" in low))
    )


def _is_auth_fail(status: int, body: str) -> bool:
    if status < 400:
        return False
    low = (body or "").lower()
    return "authentication credentials" in low or "linkedin authentication" in low


def _people_search(keywords: str, count: int = 10) -> tuple[list[dict], str | None]:
    """Use V1 /search/people — V2 RSC often returns empty for some accounts/regions."""
    try:
        r = httpx.post(
            f"{_BASE}/search/people",
            headers=_headers(),
            json={"keywords": keywords, "limit": count},
            timeout=45,
        )
    except httpx.RequestError as e:
        return [], f"request error: {e}"
    if _is_rate_limit(r.status_code, r.text):
        return [], "RATE_LIMIT"
    if _is_auth_fail(r.status_code, r.text):
        return [], "AUTH_FAIL"
    if r.status_code >= 400:
        return [], f"HTTP {r.status_code}: {r.text[:200]}"
    data = r.json() if r.content else {}
    people = data.get("people", data.get("results", []))
    results = []
    for p in people:
        url = (p.get("profileUrl") or "").strip()
        if not url:
            pid = (p.get("profileId") or p.get("publicIdentifier") or "").strip()
            if pid and "/" not in pid:
                url = f"https://www.linkedin.com/in/{pid}"
        results.append({
            "name": f"{p.get('firstName', '')} {p.get('lastName', '')}".strip(),
            "headline": p.get("headline", "") or "",
            "linkedin_profile": url,
            "current_company": (
                p.get("currentCompany")
                or p.get("currentPosition")
                or ""
            ),
        })
    return results, None


def _confirm_profile(url: str, stated_name: str, company: str = "") -> tuple[bool, str | None]:
    """Optional second check via profile API. Returns (ok, error)."""
    m = re.search(r"linkedin\.com/in/([^/?#]+)", url or "", flags=re.I)
    if not m:
        return False, "bad url"
    slug = m.group(1).strip("/")
    try:
        r = httpx.get(
            f"{_BASE}/profile",
            headers=_headers(),
            params={"profileId": slug},
            timeout=15,
        )
    except httpx.RequestError as e:
        return False, f"request error: {e}"
    if _is_rate_limit(r.status_code, r.text):
        return False, "RATE_LIMIT"
    if r.status_code == 404:
        return False, "profile not found"
    if r.status_code >= 400:
        return False, f"HTTP {r.status_code}: {r.text[:200]}"
    data = r.json() if r.content else {}
    profile = data.get("profile") if isinstance(data.get("profile"), dict) else data
    real = f"{profile.get('firstName', '')} {profile.get('lastName', '')}".strip()
    if not real:
        return False, "empty profile name"
    if not _name_matches(stated_name, real):
        return False, f"name mismatch: {real!r}"
    headline = profile.get("headline", "") or ""
    if company and not _company_mentioned(company, headline):
        return False, f"headline lacks company: {headline!r}"
    return True, None


def _pick_match(person_name: str, company: str, hits: list[dict]) -> str | None:
    for h in hits:
        url = (h.get("linkedin_profile") or "").strip()
        if not url or "linkedin.com/in/" not in url.lower():
            continue
        if not _name_matches(person_name, h.get("name") or ""):
            continue
        if not _company_mentioned(company, h.get("headline") or "", h.get("current_company") or ""):
            continue
        return url.rstrip("/")
    return None


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--sheet", required=True)
    ap.add_argument("--output", required=True)
    ap.add_argument("--start", type=int, default=0, help="0-indexed data row to start at")
    ap.add_argument("--limit", type=int, default=0, help="0 = all remaining")
    ap.add_argument("--sleep", type=float, default=10.0,
                    help="Seconds between ConnectSafely people_search calls (default 10)")
    ap.add_argument(
        "--confirm-profile",
        action="store_true",
        help="Also call profile API to re-check the chosen URL (avoid — uses LinkedIn profile quota)",
    )
    ap.add_argument(
        "--batch-size",
        type=int,
        default=0,
        help="If >0 with --limit unset, only process this many data rows then exit (for manual batches)",
    )
    args = ap.parse_args()

    out_path = Path(args.output)
    src = out_path if out_path.exists() else Path(args.input)
    wb = openpyxl.load_workbook(src)
    ws = wb[args.sheet]

    header = [c.value for c in ws[1]]
    col = {name: i + 1 for i, name in enumerate(header) if name}
    if "team" not in col or "name" not in col:
        sys.exit("Sheet missing 'name' or 'team' column")

    data_rows = [r for r in range(2, ws.max_row + 1) if ws.cell(r, col["name"]).value]
    limit = args.limit or args.batch_size or 0
    batch = data_rows[args.start: args.start + limit] if limit else data_rows[args.start:]

    searched = filled = no_hit = skipped = errors = 0
    print(
        f"Refilling missing team LinkedIn for {len(batch)} rows "
        f"(start={args.start}, sleep={args.sleep}s, confirm_profile={args.confirm_profile})...",
        file=sys.stderr,
    )

    for i, r in enumerate(batch, 1):
        company = str(ws.cell(r, col["name"]).value).strip()
        cell = ws.cell(r, col["team"]).value
        people = _parse_team(cell)
        if not people:
            continue

        changed = False
        for person in people:
            name, title, bio, linkedin = person
            if not name:
                continue
            if linkedin and "linkedin.com/in/" in linkedin.lower():
                skipped += 1
                continue

            keywords = f'"{name}" {company}'
            hits, err = _people_search(keywords)
            searched += 1
            if err == "RATE_LIMIT":
                print(
                    f"\nRATE_LIMIT hit at row {r} ({company}) — stopping cleanly.",
                    file=sys.stderr,
                    flush=True,
                )
                print(f"Resume with: --start {args.start + i - 1}", file=sys.stderr)
                wb.save(out_path)
                print(
                    f"Saved {out_path} (searched={searched} filled={filled} "
                    f"no_hit={no_hit} errors={errors})",
                    file=sys.stderr,
                )
                return
            if err == "AUTH_FAIL":
                print(
                    f"\nConnectSafely LinkedIn AUTH_FAIL at row {r} ({company}) — stopping cleanly.",
                    file=sys.stderr,
                    flush=True,
                )
                print(f"Resume with: --start {args.start + i - 1}", file=sys.stderr)
                wb.save(out_path)
                return
            if err:
                errors += 1
                print(f"  ! row {r} {name!r}: {err}", file=sys.stderr, flush=True)
                time.sleep(args.sleep)
                continue

            url = _pick_match(name, company, hits)
            if not url:
                no_hit += 1
                time.sleep(args.sleep)
                continue

            if args.confirm_profile:
                ok, cerr = _confirm_profile(url, name, company)
                if cerr == "RATE_LIMIT":
                    print(
                        f"\nRATE_LIMIT (profile) at row {r} ({company}) — stopping cleanly.",
                        file=sys.stderr,
                        flush=True,
                    )
                    print(f"Resume with: --start {args.start + i - 1}", file=sys.stderr)
                    wb.save(out_path)
                    return
                if cerr or not ok:
                    no_hit += 1
                    print(
                        f"  ~ row {r}: {name!r} search hit rejected ({cerr or 'name mismatch'})",
                        file=sys.stderr,
                        flush=True,
                    )
                    time.sleep(args.sleep)
                    continue

            person[3] = url
            filled += 1
            changed = True
            print(f"  ✓ row {r}: {name!r} @ {company} -> {url}", file=sys.stderr, flush=True)
            time.sleep(args.sleep)

        if changed:
            ws.cell(r, col["team"]).value = _render_team(people)
            wb.save(out_path)

        if i % 20 == 0:
            print(
                f"[{i}/{len(batch)}] searched={searched} filled={filled} "
                f"no_hit={no_hit} errors={errors}",
                file=sys.stderr,
                flush=True,
            )

    wb.save(out_path)
    print(
        f"\nDone. searched={searched} filled={filled} no_hit={no_hit} "
        f"skipped_has_url={skipped} errors={errors}",
        file=sys.stderr,
    )
    print(f"Saved {out_path}", file=sys.stderr)


if __name__ == "__main__":
    main()
