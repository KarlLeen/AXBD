"""Push W6 listing spreadsheet rows to AthenaX via Internal Service API.

Uses AthenaXClient (ATHENAX_API_URL + INTERNAL_API_KEY from .env).
Creates PENDING products (imported=true). Skips rows that already exist by name.

Usage:
    # dry-run
    uv run python scripts/push_listing_to_athenax.py \\
        --input data/athenax_listing_w6_team_work.xlsx \\
        --sheet "Week 6 (Huge list)" --dry-run --limit 5

    # real push
    uv run python scripts/push_listing_to_athenax.py \\
        --input data/athenax_listing_w6_team_work.xlsx \\
        --sheet "Week 6 (Huge list)" --limit 5
"""
from __future__ import annotations

import argparse
import re
import sys
import time
from pathlib import Path

_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(_ROOT))
from dotenv import load_dotenv

load_dotenv(_ROOT / ".env")

import openpyxl

from athenax.api.athenax_client import AthenaXClient, _map_stage, _valid_url


def _clean(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in {"not found", "none", "n/a", "nan"} else s


def _parse_team(cell: str) -> list[dict]:
    if not cell:
        return []
    out = []
    for chunk in str(cell).split(" ; "):
        parts = [p.strip() for p in chunk.split(" | ")]
        while len(parts) < 4:
            parts.append("")
        name, title, bio, linkedin = parts[:4]
        if not name:
            continue
        # Cap bio for API (max 300)
        bio = bio[:300] if bio else ""
        # Prefer bullet-less single-line for short notes
        if bio and len(bio) > 300:
            bio = bio[:297] + "..."
        m = {
            "name": name[:100],
            "title": (title[:150] if title else None),
            "bio": bio or None,
            "linkedin": linkedin if linkedin and "linkedin.com/in/" in linkedin.lower() else None,
            "twitter": None,
        }
        # 5th field twitter if present
        if len(parts) > 4 and parts[4]:
            m["twitter"] = parts[4]
        out.append(m)
    return out


def _parse_backers(cell: str) -> list[str]:
    if not cell:
        return []
    return [b.strip() for b in str(cell).split(" ; ") if b.strip()]


def _row_to_lead(row: dict) -> dict:
    """Map spreadsheet columns → AthenaXClient.push_product lead dict."""
    stage_raw = _clean(row.get("stage"))
    # Not Active → leave stage empty (do not map / do not skip the product)
    if stage_raw.lower() == "not active":
        stage = None
    else:
        stage = stage_raw or None

    founded = _clean(row.get("founded"))
    team = _parse_team(_clean(row.get("team")))
    backers = _parse_backers(_clean(row.get("backers")))

    return {
        "name": _clean(row.get("name")),
        "website": _valid_url(row.get("Website") or row.get("website")),
        "url": _valid_url(row.get("Website") or row.get("website")),
        "short_description": _clean(row.get("short_desc"))[:150] or None,
        "description": _clean(row.get("description")) or None,
        "stage": stage,
        "founded": founded or None,
        "sector": _clean(row.get("category")) or None,
        "subcategory": _clean(row.get("subcategory")) or None,
        "twitter_url": _valid_url(row.get("Twitter") or row.get("twitter")),
        "github_url": _valid_url(row.get("GitHub") or row.get("github")),
        "discord_url": _valid_url(row.get("Discord") or row.get("discord")),
        "docs_url": _valid_url(row.get("Docs") or row.get("docs")),
        "team": team,
        "backers": backers,
    }


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--sheet", required=True)
    ap.add_argument("--start", type=int, default=0, help="0-indexed data row")
    ap.add_argument("--limit", type=int, default=0, help="0 = all remaining")
    ap.add_argument("--sleep", type=float, default=1.0)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.input, read_only=True, data_only=True)
    ws = wb[args.sheet]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = {h: i for i, h in enumerate(headers) if h}

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        name = r[col_idx["name"]] if "name" in col_idx else None
        if not name:
            continue
        rows.append({h: r[i] for h, i in col_idx.items()})
    wb.close()

    batch = rows[args.start: args.start + args.limit] if args.limit else rows[args.start:]
    client = AthenaXClient()
    print(
        f"{'[DRY-RUN] ' if args.dry_run else ''}Pushing {len(batch)} rows "
        f"→ {client.base_url} (start={args.start})",
        flush=True,
    )

    created = skipped = failed = 0
    for i, row in enumerate(batch, 1):
        lead = _row_to_lead(row)
        name = lead["name"]
        if not name:
            continue

        # Resolve category (retry — transient DNS/SSL made this look "not found")
        category_id = None
        if lead.get("sector"):
            for attempt in range(3):
                category_id = client.resolve_category_id(lead["sector"])
                if category_id is not None:
                    break
                time.sleep(1.5 * (attempt + 1))
            if category_id is None:
                print(f"  ! {name}: category {lead['sector']!r} not found — "
                      f"dropping subcategory to avoid 400", flush=True)
                # API requires parent category when otherSubcategoryName is set
                lead["subcategory"] = None

        # Dedup (retry transient network)
        existing = None
        lookup_err = None
        for attempt in range(3):
            try:
                existing = client.get_product_by_name(name)
                lookup_err = None
                break
            except Exception as e:
                lookup_err = e
                time.sleep(1.5 * (attempt + 1))
        if lookup_err is not None:
            print(f"  ✗ [{i}/{len(batch)}] {name}: lookup failed: {lookup_err}", flush=True)
            failed += 1
            time.sleep(args.sleep)
            continue

        if existing:
            print(
                f"  ↩ [{i}/{len(batch)}] skip existing {name!r} "
                f"(id={existing.get('id')} status={existing.get('status')})",
                flush=True,
            )
            skipped += 1
            time.sleep(args.sleep)
            continue

        if args.dry_run:
            team_n = len(lead.get("team") or [])
            backers_n = len(lead.get("backers") or [])
            print(
                f"  · [{i}/{len(batch)}] would create {name!r} "
                f"cat={lead.get('sector')}→{category_id} stage={lead.get('stage')!r} "
                f"team={team_n} backers={backers_n} url={lead.get('url')}",
                flush=True,
            )
            created += 1
            continue

        push_err = None
        pid = None
        for attempt in range(3):
            try:
                pid = client.push_product(lead, evaluation={}, category_id=category_id)
                push_err = None
                break
            except Exception as e:
                push_err = e
                # Don't retry hard 400s (except we already stripped subcategory)
                resp = getattr(e, "response", None)
                if resp is not None and resp.status_code == 400:
                    break
                time.sleep(2.0 * (attempt + 1))

        if push_err is None and pid:
            print(
                f"  ✓ [{i}/{len(batch)}] created PENDING {name!r} → id={pid}",
                flush=True,
            )
            created += 1
        else:
            detail = ""
            resp = getattr(push_err, "response", None) if push_err else None
            if resp is not None:
                detail = f" body={resp.text[:300]}"
            print(f"  ✗ [{i}/{len(batch)}] {name!r}: {push_err}{detail}", flush=True)
            failed += 1

        time.sleep(args.sleep)

    print(
        f"\nDone. created={created} skipped_existing={skipped} failed={failed}"
        f"{' (dry-run)' if args.dry_run else ''}",
        flush=True,
    )


if __name__ == "__main__":
    main()
