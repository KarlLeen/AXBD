"""Deterministic (non-LLM) cleanup pass for the 'team' column's LinkedIn links.

The enrichment crew occasionally attaches the wrong person's LinkedIn profile to a
team member (matched on a similar first name, wrong last name/company — e.g. Loom's
"Shahed Khan" got linked to an unrelated "Shahedul Huq Khandkar" at McKesson Canada).
Rather than trust the LLM's judgment on identity, this fetches the REAL name behind
each linkedin URL via ConnectSafely's profile endpoint and does a plain string
comparison against the name already on the row — no model involved.

Crash-safe / resumable: saves after every row, skip already-checked rows on rerun
unless --force. Stops cleanly (not a crash) if ConnectSafely's daily profile-view
limit is hit, so a later run can pick up where this one left off.

Usage:
    uv run python scripts/verify_team_linkedin.py \\
        --input data/athenax_listing_w6_enriched.xlsx \\
        --sheet "Week 6 (Huge list)" \\
        --output data/athenax_listing_w6_enriched.xlsx
"""
import argparse
import difflib
import os
import re
import sys
import time
import unicodedata
from pathlib import Path

_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(_ROOT))
from dotenv import load_dotenv
load_dotenv(_ROOT / ".env")

import httpx
import openpyxl

_BASE = "https://api.connectsafely.ai/linkedin"


def _headers() -> dict:
    key = os.getenv("CONNECTSAFELY_API_KEY", "")
    if not key:
        sys.exit("CONNECTSAFELY_API_KEY not set")
    return {"Authorization": f"Bearer {key}", "Content-Type": "application/json"}


def _slug(url: str) -> str | None:
    m = re.search(r"linkedin\.com/in/([^/?#]+)", url or "", flags=re.I)
    return m.group(1).strip("/") if m else None


def _profile_payload(data: dict) -> dict:
    """ConnectSafely nests person fields under `profile`; older shapes were flat."""
    if isinstance(data.get("profile"), dict):
        return data["profile"]
    return data


def _real_profile(slug: str) -> tuple[dict | None, str | None]:
    """Returns (profile_dict, error). error is a short reason string, or None on success."""
    try:
        r = httpx.get(f"{_BASE}/profile", headers=_headers(),
                       params={"profileId": slug}, timeout=15)
    except httpx.RequestError as e:
        return None, f"request error: {e}"
    if r.status_code == 404:
        return None, "profile not found (dead/invalid slug)"
    if r.status_code == 429:
        return None, "RATE_LIMIT"
    if r.status_code >= 400:
        body = r.text[:200]
        low = body.lower()
        if (
            r.status_code == 429
            or "rate limit" in low
            or "rate_limit" in low
            or "daily limit" in low
            or "quota exceeded" in low
            or "too many requests" in low
            or ("limit" in low and ("exceed" in low or "reached" in low or "month" in low))
        ):
            return None, "RATE_LIMIT"
        if "authentication credentials" in low or "linkedin authentication" in low:
            return None, "AUTH_FAIL"
        return None, f"HTTP {r.status_code}: {body}"
    data = r.json() if r.content else {}
    if data.get("success") is False:
        return None, f"api error: {data.get('error') or data}"
    profile = _profile_payload(data if isinstance(data, dict) else {})
    name = f"{profile.get('firstName', '')} {profile.get('lastName', '')}".strip()
    if not name:
        return None, "empty profile name"
    return profile, None


def _fold(s: str) -> str:
    s = unicodedata.normalize("NFKD", s or "")
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z ]", "", s.lower())


def _name_matches(stated: str, real: str | None) -> bool:
    if not stated or not real:
        return False
    a = _fold(stated).split()
    b = _fold(real).split()
    if not a or not b:
        return False
    # Last-token (surname) must match closely; first name allows nicknames/typos.
    if difflib.SequenceMatcher(None, a[-1], b[-1]).ratio() < 0.8:
        return False
    return difflib.SequenceMatcher(None, " ".join(a), " ".join(b)).ratio() >= 0.55


def _company_mentioned(company: str, *fields: str) -> bool:
    blob = " ".join(f or "" for f in fields).lower()
    if not blob:
        return False
    raw = re.sub(r"\([^)]*\)", " ", company).strip()
    tokens = [t for t in re.sub(r"[^a-z0-9 ]", " ", raw.lower()).split() if len(t) >= 3]
    stop = {"the", "and", "inc", "ltd", "llc", "labs", "lab", "ai", "web", "app", "for"}
    tokens = [t for t in tokens if t not in stop]
    if not tokens:
        needle = re.sub(r"[^a-z0-9]", "", raw.lower())
        return len(needle) >= 3 and needle in re.sub(r"[^a-z0-9]", "", blob)
    tokens.sort(key=len, reverse=True)
    return any(t in blob for t in tokens[:2])


_GARBAGE_HEADLINES = {
    "", "about this member", "withdraw invitation", "linkedin member", "null", "none",
}


def _usable_headline(headline: str) -> bool:
    h = (headline or "").strip().lower()
    if h in _GARBAGE_HEADLINES:
        return False
    if len(h) < 8:
        return False
    return True


def _parse_team(cell: str) -> list[list[str]]:
    if not cell:
        return []
    people = []
    for chunk in cell.split(" ; "):
        parts = chunk.split(" | ")
        while len(parts) < 4:
            parts.append("")
        people.append(parts[:4])  # name, title, bio, linkedin
    return people


def _render_team(people: list[list[str]]) -> str:
    out = []
    for name, title, bio, linkedin in people:
        parts = [p for p in (name, title, bio, linkedin) if p]
        out.append(" | ".join(parts))
    return " ; ".join(out)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--sheet", required=True)
    ap.add_argument("--output", required=True)
    ap.add_argument("--start", type=int, default=0, help="0-indexed data row to start at")
    ap.add_argument("--limit", type=int, default=0, help="0 = all remaining")
    ap.add_argument("--sleep", type=float, default=1.0)
    args = ap.parse_args()

    out_path = Path(args.output)
    src = out_path if out_path.exists() else Path(args.input)
    wb = openpyxl.load_workbook(src)
    ws = wb[args.sheet]

    header = [c.value for c in ws[1]]
    col = {name: i + 1 for i, name in enumerate(header) if name}
    if "team" not in col or "name" not in col:
        sys.exit("Sheet missing 'name' or 'team' column")

    data_rows = [r for r in range(2, ws.max_row + 1) if ws.cell(r, col["name"]).value]
    batch = data_rows[args.start: args.start + args.limit] if args.limit else data_rows[args.start:]

    checked = kept = stripped = skipped_no_url = errors = 0
    print(f"Verifying team LinkedIn links for {len(batch)} rows...", file=sys.stderr)

    for i, r in enumerate(batch, 1):
        company = ws.cell(r, col["name"]).value
        cell = ws.cell(r, col["team"]).value
        people = _parse_team(cell)
        if not people:
            continue

        changed = False
        for person in people:
            name, title, bio, linkedin = person
            if not linkedin or not name:
                skipped_no_url += 1
                continue
            slug = _slug(linkedin)
            if not slug:
                continue

            profile, err = _real_profile(slug)
            checked += 1

            if err == "RATE_LIMIT":
                print(f"\nRATE_LIMIT hit at row {r} ({company}) — stopping cleanly.",
                      file=sys.stderr, flush=True)
                print(f"Resume with: --start {args.start + i - 1}", file=sys.stderr)
                wb.save(out_path)
                print(f"Saved {out_path} (checked={checked} kept={kept} "
                      f"stripped={stripped} errors={errors})", file=sys.stderr)
                return
            if err == "AUTH_FAIL":
                print(f"\nConnectSafely LinkedIn AUTH_FAIL at row {r} ({company}) — stopping cleanly.",
                      file=sys.stderr, flush=True)
                print(f"Resume with: --start {args.start + i - 1}", file=sys.stderr)
                wb.save(out_path)
                print(f"Saved {out_path} (checked={checked} kept={kept} "
                      f"stripped={stripped} errors={errors})", file=sys.stderr)
                return
            if err:
                # Dead / invalid profile URLs are wrong links — strip them so refill can retry.
                if "not found" in err or "dead" in err or "empty" in err:
                    print(f"  ✗ row {r}: {name!r} -> {linkedin} ({err}) — stripping",
                          file=sys.stderr, flush=True)
                    person[3] = ""
                    stripped += 1
                    changed = True
                else:
                    errors += 1
                    print(f"  ! row {r} {name!r}: {err}", file=sys.stderr, flush=True)
                time.sleep(args.sleep)
                continue

            real_name = f"{profile.get('firstName', '')} {profile.get('lastName', '')}".strip()
            headline = profile.get("headline", "") or ""
            if not _name_matches(name, real_name):
                print(f"  ✗ row {r}: {name!r} -> {linkedin} is actually {real_name!r} — stripping",
                      file=sys.stderr, flush=True)
                person[3] = ""
                stripped += 1
                changed = True
            elif _company_mentioned(company, headline):
                kept += 1
            elif not _usable_headline(headline):
                # Privacy-garbled headline — name matched; keep, don't over-strip.
                kept += 1
            else:
                # Substantive headline with no company mention → likely a homonym.
                print(f"  ✗ row {r}: {name!r} -> {linkedin} name ok ({real_name!r}) but "
                      f"headline {headline!r} lacks {company!r} — stripping",
                      file=sys.stderr, flush=True)
                person[3] = ""
                stripped += 1
                changed = True

            time.sleep(args.sleep)

        if changed:
            ws.cell(r, col["team"]).value = _render_team(people)
            wb.save(out_path)

        if i % 20 == 0:
            print(f"[{i}/{len(batch)}] checked={checked} kept={kept} stripped={stripped}",
                  file=sys.stderr, flush=True)

    wb.save(out_path)
    print(f"\nDone. checked={checked} kept={kept} stripped={stripped} "
          f"skipped_no_url={skipped_no_url} errors={errors}", file=sys.stderr)
    print(f"Saved {out_path}", file=sys.stderr)


if __name__ == "__main__":
    main()
