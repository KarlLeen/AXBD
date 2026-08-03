"""
Backfill an AthenaX listing spreadsheet (name/website/twitter already filled,
everything else blank) using the AthenaX Partnership Agent's research tools
(GitHub, Serper web search, LinkedIn/ConnectSafely, CoinGecko — all via .env).

Runs ONE enrichment crew per company and writes:
  • progress to stderr
  • a JSON progress file (data/enrich_progress.json by default) for the dashboard
  • the filled .xlsx after every row (crash-safe / resumable)

Usage:
    uv run python scripts/enrich_listing.py \\
        --input data/athenax_listing_w6.xlsx \\
        --sheet "Week 6 (Huge list)" \\
        --start 0 --limit 0 \\
        --output data/athenax_listing_w6_enriched.xlsx
"""
import argparse
import os
import sys
import time
from pathlib import Path

_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(_ROOT))
from dotenv import load_dotenv
load_dotenv(_ROOT / ".env")

import openpyxl

from athenax.main import _balanced_extract, _try_parse
from athenax.crew import build_enrichment_crew
from athenax.enrich_progress import (
    DEFAULT_PATH as PROGRESS_PATH,
    append_recent,
    default_state,
    load as load_progress,
    save as save_progress,
)

TARGET_COLUMNS = [
    "category", "subcategory", "short_desc", "description", "stage", "founded",
    "Discord", "GitHub", "Docs", "team", "backers",
]

FIELD_TO_HEADER = {
    "category": "category",
    "subcategory": "subcategory",
    "short_desc": "short_desc",
    "description": "description",
    "stage": "stage",
    "founded": "founded",
    "discord_url": "Discord",
    "github_url": "GitHub",
    "docs_url": "Docs",
}


def _clean(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() == "not found" else s


def _team_cell(team) -> str:
    if not isinstance(team, list):
        return ""
    out = []
    for m in team:
        if not isinstance(m, dict):
            continue
        parts = [_clean(m.get(k)) for k in ("name", "title", "bio", "linkedin")]
        if parts[0]:
            out.append(" | ".join(parts))
    return " ; ".join(out)


def _backers_cell(backers) -> str:
    if not isinstance(backers, list):
        return ""
    return " ; ".join(_clean(b) for b in backers if _clean(b))


def _extract_object(text: str) -> dict | None:
    """Pull the first balanced {...} object out of a crew's raw text output."""
    import re
    text = re.sub(r"```(?:json)?\s*", "", text).strip().rstrip("`").strip()
    pos = 0
    while True:
        pos = text.find("{", pos)
        if pos == -1:
            return None
        chunk = _balanced_extract(text, pos)
        if chunk:
            parsed = _try_parse(chunk)
            if isinstance(parsed, dict):
                return parsed
            pos += len(chunk)
        else:
            pos += 1


def _count_filled(ws, col: dict, data_rows: list[int]) -> int:
    return sum(1 for r in data_rows if _clean(ws.cell(r, col["category"]).value))


def _needs_enrich(ws, col: dict, r: int, force: bool) -> bool:
    """Skip only when category is present AND at least one of founded/team/backers is set.

    Sparse rows (category filled but no team/year/investors) get another pass — useful
    after upgrading tools (e.g. wiring Twitter) without --force on every row.
    """
    if force:
        return True
    if not _clean(ws.cell(r, col["category"]).value):
        return True
    for key in ("founded", "team", "backers"):
        if _clean(ws.cell(r, col[key]).value):
            return False
    return True


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, help="Path to the source .xlsx")
    ap.add_argument("--sheet", required=True, help="Sheet name to enrich")
    ap.add_argument("--start", type=int, default=0, help="0-indexed data row to start at")
    ap.add_argument("--limit", type=int, default=0,
                    help="Number of rows to enrich (0 = all remaining from --start)")
    ap.add_argument("--output", required=True, help="Path to write the filled .xlsx")
    ap.add_argument("--force", action="store_true",
                    help="Re-enrich rows that already have a category filled in "
                         "(default: skip them, so a batch is safely resumable/idempotent)")
    ap.add_argument("--retries", type=int, default=2,
                    help="Attempts per company before giving up (handles transient "
                         "max-iteration / LLM failures)")
    ap.add_argument("--progress", default=str(PROGRESS_PATH),
                    help="JSON progress file the dashboard polls")
    ap.add_argument("--sleep", type=float, default=2.0,
                    help="Seconds to pause between companies (rate-limit friendly)")
    ap.add_argument("--fields", default="",
                    help="Comma-separated subset of columns to overwrite (e.g. 'team' or "
                         "'team,backers'). Default: all columns. The crew still researches "
                         "everything per company either way — this only restricts what gets "
                         "written, so already-good columns can't regress from a targeted rerun.")
    args = ap.parse_args()
    write_fields = {f.strip() for f in args.fields.split(",") if f.strip()} or None

    progress_path = Path(args.progress)
    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # Prefer the working output on resume so earlier fills aren't lost when
    # --input is still the pristine source spreadsheet.
    src = out_path if out_path.exists() else Path(args.input)
    wb = openpyxl.load_workbook(src)
    ws = wb[args.sheet]

    header = [c.value for c in ws[1]]
    col = {name: i + 1 for i, name in enumerate(header) if name}
    required = {"name", "Website", "Twitter", *TARGET_COLUMNS}
    missing = required - set(col)
    if missing:
        sys.exit(f"Sheet is missing expected columns: {sorted(missing)}")

    data_rows = [r for r in range(2, ws.max_row + 1) if ws.cell(r, col["name"]).value]
    if args.limit and args.limit > 0:
        batch = data_rows[args.start: args.start + args.limit]
    else:
        batch = data_rows[args.start:]

    already = _count_filled(ws, col, data_rows)
    print(
        f"Enriching {len(batch)} rows (start={args.start}, limit={args.limit or 'all'}) "
        f"of {len(data_rows)} total data rows "
        f"({already} already have category)...",
        file=sys.stderr,
    )

    from datetime import datetime, timezone
    state = default_state()
    state.update({
        "running": True,
        "pid": os.getpid(),
        "started_at": datetime.now(timezone.utc).isoformat(),
        "input": str(Path(args.input).resolve()),
        "output": str(out_path.resolve()),
        "sheet": args.sheet,
        "total": len(data_rows),
        "done": already,
        "skipped": 0,
        "failed": 0,
        "last_status": "running",
        "last_error": None,
        "recent": load_progress(progress_path).get("recent") or [],
    })
    save_progress(state, progress_path)

    try:
        for i, r in enumerate(batch, 1):
            name = ws.cell(r, col["name"]).value
            website = ws.cell(r, col["Website"]).value
            twitter = ws.cell(r, col["Twitter"]).value

            state["current_index"] = args.start + i
            state["current_name"] = str(name)
            state["current_row"] = r
            save_progress(state, progress_path)

            if not _needs_enrich(ws, col, r, args.force):
                print(f"[{i}/{len(batch)}] row {r}: {name!r} — already filled, skipping",
                      file=sys.stderr, flush=True)
                state["skipped"] += 1
                append_recent(state, {
                    "name": str(name), "status": "skipped", "seconds": 0,
                    "at": datetime.now(timezone.utc).isoformat(),
                })
                save_progress(state, progress_path)
                continue

            print(f"[{i}/{len(batch)}] row {r}: {name!r} ...", file=sys.stderr, flush=True)

            wrote_any = False
            last_err = None
            elapsed = 0.0
            for attempt in range(1, args.retries + 1):
                t0 = time.time()
                try:
                    crew = build_enrichment_crew({
                        "name": name, "website": website, "twitter": twitter,
                    })
                    result = crew.kickoff()
                    data = _extract_object(str(result.raw if hasattr(result, "raw") else result))
                except Exception as e:
                    last_err = str(e)
                    print(f"  ! attempt {attempt} error for {name!r}: {e}",
                          file=sys.stderr, flush=True)
                    continue

                if not data:
                    last_err = "could not parse JSON"
                    print(f"  ! attempt {attempt}: could not parse JSON for {name!r}",
                          file=sys.stderr, flush=True)
                    continue

                for field, header_name in FIELD_TO_HEADER.items():
                    if write_fields is not None and field not in write_fields:
                        continue
                    val = _clean(data.get(field))
                    if val:
                        ws.cell(r, col[header_name]).value = val
                        wrote_any = True

                if write_fields is None or "team" in write_fields:
                    team_val = _team_cell(data.get("team"))
                    if team_val:
                        ws.cell(r, col["team"]).value = team_val
                        wrote_any = True

                if write_fields is None or "backers" in write_fields:
                    backers_val = _backers_cell(data.get("backers"))
                    if backers_val:
                        ws.cell(r, col["backers"]).value = backers_val
                        wrote_any = True

                elapsed = time.time() - t0
                if wrote_any:
                    print(f"  done in {elapsed:.0f}s (attempt {attempt})",
                          file=sys.stderr, flush=True)
                    break
                last_err = f"no usable fields (keys: {sorted(data.keys())})"
                print(f"  ! attempt {attempt}: parsed JSON had no usable fields for {name!r} "
                      f"(keys: {sorted(data.keys())})", file=sys.stderr, flush=True)

            if wrote_any:
                state["done"] = _count_filled(ws, col, data_rows)
                append_recent(state, {
                    "name": str(name), "status": "ok", "seconds": round(elapsed),
                    "at": datetime.now(timezone.utc).isoformat(),
                })
            else:
                state["failed"] += 1
                print(f"  ✗ gave up on {name!r} after {args.retries} attempts — left blank",
                      file=sys.stderr, flush=True)
                append_recent(state, {
                    "name": str(name), "status": "failed", "seconds": round(elapsed),
                    "error": last_err, "at": datetime.now(timezone.utc).isoformat(),
                })

            wb.save(out_path)
            save_progress(state, progress_path)

            if args.sleep > 0 and i < len(batch):
                time.sleep(args.sleep)

        state["last_status"] = "ok"
        state["current_name"] = None
        state["current_row"] = None
    except Exception as exc:
        state["last_status"] = "error"
        state["last_error"] = str(exc)
        raise
    finally:
        state["running"] = False
        state["pid"] = None
        state["finished_at"] = datetime.now(timezone.utc).isoformat()
        state["done"] = _count_filled(ws, col, data_rows)
        try:
            wb.save(out_path)
        except Exception:
            pass
        save_progress(state, progress_path)

    print(f"Saved {out_path}", file=sys.stderr)


if __name__ == "__main__":
    main()
