"""Offline unit tests for lightweight listing-scout helpers."""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pytest

from athenax.listing_scout import (
    W6_HEADERS,
    append_leads_to_xlsx,
    build_exclude_set,
    casefold_name,
    filter_scout_leads,
    format_exclude_context,
    lead_to_w6_row,
    load_inventory,
    looks_like_website,
    names_from_inventory,
    names_from_xlsx,
    normalize_lead,
    normalize_twitter,
    parse_scout_json,
    sector_hint_for_round,
)


def test_parse_scout_json_plain_array():
    text = json.dumps(
        [
            {
                "name": "Acme Agents",
                "website": "https://acmeagents.example.io",
                "twitter": "@acmeagents",
                "category": "AI & Agents",
            },
            {
                "name": "ChainKit",
                "website": "https://chainkit.dev",
                "twitter": "https://x.com/chainkit",
            },
        ]
    )
    leads = parse_scout_json(text)
    assert len(leads) == 2
    assert leads[0]["name"] == "Acme Agents"


def test_parse_scout_json_fenced_and_prose():
    text = """
Here are the leads I found:

```json
[
  {"name": "Nova RWA", "website": "https://novarwa.com", "twitter": ""},
  {"name": "BotFarm", "website": "https://botfarm.ai", "twitter": "@botfarm"}
]
```

Hope this helps!
"""
    leads = parse_scout_json(text)
    assert len(leads) == 2
    assert leads[1]["name"] == "BotFarm"


def test_parse_scout_json_trailing_comma_repair():
    text = '[{"name": "X", "website": "https://x.example.com", "twitter": "",},]'
    leads = parse_scout_json(text)
    assert len(leads) == 1
    assert leads[0]["name"] == "X"


def test_parse_scout_json_empty():
    assert parse_scout_json("") == []
    assert parse_scout_json("no json here") == []


def test_looks_like_website():
    assert looks_like_website("https://foo.bar/path")
    assert looks_like_website("http://foo.io")
    assert not looks_like_website("not-a-url")
    assert not looks_like_website("https://example.com")
    assert not looks_like_website("")
    assert not looks_like_website(None)


def test_normalize_twitter():
    assert normalize_twitter("@foo_bar") == "https://x.com/foo_bar"
    assert normalize_twitter("https://twitter.com/foo_bar") == "https://x.com/foo_bar"
    assert normalize_twitter("https://x.com/foo_bar/status/1") == "https://x.com/foo_bar"
    assert normalize_twitter("not found") == ""
    assert normalize_twitter("") == ""


def test_normalize_lead_requires_website():
    assert normalize_lead({"name": "A", "website": "nope"}) is None
    assert normalize_lead({"name": "", "website": "https://a.io"}) is None
    n = normalize_lead(
        {"name": "A", "Website": "https://a.io/", "Twitter": "@a", "category": "crypto"}
    )
    assert n == {
        "name": "A",
        "website": "https://a.io",
        "twitter": "https://x.com/a",
        "category": "Crypto",
    }


def test_filter_scout_leads_exclude_and_dedup():
    excluded = {"acme agents", "oldco"}
    raw = [
        {"name": "Acme Agents", "website": "https://acme.io"},
        {"name": "NewCo", "website": "https://newco.dev", "twitter": "@newco"},
        {"name": "NewCo", "website": "https://newco.dev/about"},  # dup name
        {"name": "Twin", "website": "https://newco.dev"},  # dup website after first kept
        {"name": "NoSite", "website": "ftp://bad"},
        {"name": "Other", "website": "https://other.ai", "category": "Robotics"},
    ]
    # After keeping NewCo with https://newco.dev, Twin shares same website key.
    kept, stats = filter_scout_leads(raw, excluded)
    names = [k["name"] for k in kept]
    assert "Acme Agents" not in names
    assert "NoSite" not in names
    assert "NewCo" in names
    assert "Other" in names
    assert stats["excluded"] >= 1
    assert stats["no_website"] >= 1
    assert stats["kept"] == len(kept)


def test_filter_respects_seen_sets():
    raw = [{"name": "Fresh", "website": "https://fresh.dev"}]
    kept, _ = filter_scout_leads(
        raw,
        excluded=set(),
        seen_names={"fresh"},
    )
    assert kept == []
    kept2, _ = filter_scout_leads(
        raw,
        excluded=set(),
        seen_websites={"https://fresh.dev"},
    )
    assert kept2 == []


def test_inventory_and_exclude_set(tmp_path: Path):
    inv = {
        "approved_names": ["Alpha", "Beta"],
        "pending_names": ["gamma", "Beta"],
    }
    inv_path = tmp_path / "inv.json"
    inv_path.write_text(json.dumps(inv), encoding="utf-8")

    loaded = load_inventory(inv_path)
    names = names_from_inventory(loaded)
    assert names == {"alpha", "beta", "gamma"}

    # Build a tiny existing output xlsx and merge.
    xlsx = tmp_path / "out.xlsx"
    append_leads_to_xlsx(
        xlsx,
        [{"name": "Delta", "website": "https://delta.io", "twitter": "", "category": ""}],
    )
    excluded = build_exclude_set(inv_path, output_xlsx=xlsx)
    assert "alpha" in excluded
    assert "delta" in excluded
    ctx = format_exclude_context(excluded)
    assert "alpha" in ctx
    assert "EXCLUSION LIST" in ctx


def test_xlsx_append_dedup_and_headers(tmp_path: Path):
    path = tmp_path / "batch.xlsx"
    leads = [
        {
            "name": "One",
            "website": "https://one.dev",
            "twitter": "https://x.com/one",
            "category": "Crypto",
        },
        {
            "name": "Two",
            "website": "https://two.dev",
            "twitter": "",
            "category": "",
        },
    ]
    n1 = append_leads_to_xlsx(path, leads)
    assert n1 == 2
    n2 = append_leads_to_xlsx(path, leads)  # dedup
    assert n2 == 0
    n3 = append_leads_to_xlsx(
        path,
        [{"name": "one", "website": "https://one.dev/x", "twitter": "", "category": ""}],
    )
    assert n3 == 0  # casefold name dedup

    assert names_from_xlsx(path) == {"one", "two"}

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header == W6_HEADERS
    assert ws.cell(2, 1).value == "One"
    assert ws.cell(2, 4).value == "https://one.dev"  # Website
    assert ws.cell(2, 5).value == "https://x.com/one"  # Twitter
    # Other enrich columns blank
    assert ws.cell(2, 6).value in ("", None)
    wb.close()


def test_lead_to_w6_row_shape():
    row = lead_to_w6_row(
        {"name": "Z", "website": "https://z.io", "twitter": "", "category": "RWA"}
    )
    assert len(row) == len(W6_HEADERS)
    assert row[0] == "Z"
    assert row[1] == "RWA"
    assert row[3] == "https://z.io"


def test_sector_hint_rotation():
    assert sector_hint_for_round(0) == "AI & Agents"
    assert sector_hint_for_round(1) == "Crypto"
    assert sector_hint_for_round(7) == "AI & Agents"
    assert sector_hint_for_round(0, override="Biotech") == "Biotech"


def test_casefold_name():
    assert casefold_name("  Foo BAR ") == "foo bar"
    assert casefold_name(None) == ""


def test_repo_inventory_loads():
    """Smoke: committed inventory is readable and has the expected keys."""
    path = Path("data/admin_products_inventory.json")
    if not path.exists():
        pytest.skip("inventory not present")
    inv = load_inventory(path)
    names = names_from_inventory(inv)
    assert len(names) > 100
    assert "approved_names" in inv
    assert "pending_names" in inv
